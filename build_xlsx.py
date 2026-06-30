#!/usr/bin/env python3
"""Build a 2-sheet Excel workbook from the two course plans (Vietnamese)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- palette ---
NAVY = "1F3864"
BLUE = "2E5496"
LIGHT = "D9E1F2"
GREY = "F2F2F2"
WARN = "FCE4D6"
WHITE = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")

wb = Workbook()


def title(ws, text, ncols, row):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=14, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row].height = 28
    return row + 1


def section(ws, text, ncols, row):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=11, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BLUE)
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[row].height = 22
    return row + 1


def banner(ws, text, ncols, row, color=WARN):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=10, color="843C0C")
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = WRAP
    ws.row_dimensions[row].height = max(30, (text.count("\n") + 1) * 15 + 8)
    for col in range(1, ncols + 1):
        ws.cell(row=row, column=col).border = BORDER
    return row + 1


def header_row(ws, headers, row):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = WRAP_CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 20
    return row + 1


def data_row(ws, values, row, zebra=False):
    maxlines = 1
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.alignment = WRAP
        c.border = BORDER
        if zebra:
            c.fill = PatternFill("solid", fgColor=GREY)
        maxlines = max(maxlines, str(v).count("\n") + 1)
    ws.row_dimensions[row].height = max(18, maxlines * 15 + 4)
    return row + 1


def kv(ws, key, val, row):
    a = ws.cell(row=row, column=1, value=key)
    a.font = Font(bold=True)
    a.alignment = WRAP
    a.border = BORDER
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    b = ws.cell(row=row, column=2, value=val)
    b.alignment = WRAP
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = BORDER
    ws.row_dimensions[row].height = max(18, (str(val).count("\n") + 1) * 15 + 4)
    return row + 1


def spacer(row):
    return row + 1


# ============================================================== SHEET 1
ws = wb.active
ws.title = "Khóa học AI"
widths = [10, 26, 46, 30, 30]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[chr(64 + i)].width = w

r = 1
r = title(ws, "KHÓA HỌC: ỨNG DỤNG AI VÀO CÔNG VIỆC", 5, r)
r = spacer(r)
r = section(ws, "TỔNG QUAN", 5, r)
r = kv(ws, "Đối tượng", "Nhân sự không chuyên kỹ thuật (ops, vận hành, hành chính, nội dung)", r)
r = kv(ws, "Thời lượng", "4 buổi", r)
r = kv(ws, "Mục tiêu cuối", "Mỗi học viên biến AI thành thói quen — ít nhất 1 công việc hằng tuần làm bằng AI", r)
r = spacer(r)

r = section(ws, "NGUYÊN TẮC XUYÊN SUỐT", 5, r)
r = banner(ws,
    "“Đào tạo không phải nghe cho biết, mà làm cho quen.”\n"
    "Ngay Buổi 1, mỗi học viên chọn 1 CÔNG VIỆC THẬT làm mỗi tuần và theo suốt 4 buổi: "
    "làm thử → viết prompt đúng → làm sâu hơn → lưu thành mẫu. Sau mỗi buổi đều có bài tập về nhà.",
    5, r, color=LIGHT)
r = spacer(r)

r = section(ws, "CHỈ SỐ THÀNH CÔNG (đo được cuối khóa)", 5, r)
for item in [
    "Mỗi học viên có 1 công việc cố định đang làm bằng AI",
    "Mỗi học viên lưu được ít nhất 3 prompt mẫu dùng lại được",
    "Cả nhóm có 1 thư viện prompt chung",
]:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(row=r, column=1, value="☐  " + item)
    c.alignment = WRAP
    for col in range(1, 6):
        ws.cell(row=r, column=col).border = BORDER
    r += 1
r = spacer(r)

r = section(ws, "CHƯƠNG TRÌNH 4 BUỔI", 5, r)
r = header_row(ws, ["Buổi", "Mục tiêu", "Nội dung chính", "Thực hành", "Bài tập về nhà"], r)

sessions = [
    ("Buổi 1\nHiểu đúng về AI",
     "Hiểu AI là gì, làm được gì, không làm được gì. Tránh kỳ vọng sai.",
     "1. Mở đầu bằng trải nghiệm thật (20' đầu): đưa ngay công việc của mình cho AI làm thử → tạo “wow” trước, lý thuyết sau.\n"
     "2. AI là gì? Giới thiệu nhanh: ChatGPT, Gemini, Claude.\n"
     "3. AI giúp gì: rút ngắn thời gian viết báo cáo, email, tìm/tóm tắt thông tin.\n"
     "4. Giới hạn: AI có thể “bịa” (đoán chữ, không tra cứu sự thật → sai rất tự tin); và vấn đề bảo mật.",
     "Tạo tài khoản, đăng nhập; chọn 1 công việc thật và làm thử lần đầu.",
     "Dùng AI cho công việc đó trong tuần; ghi lại chỗ tốt / chỗ sai."),
    ("Buổi 2\nKỹ năng ra lệnh (Prompting)",
     "Biết cách hỏi để AI trả lời đúng ý, không chung chung.",
     "1. Vì sao AI trả lời dở? → Câu hỏi thiếu thông tin, không phải AI kém.\n"
     "2. Công thức: Vai trò + Nhiệm vụ + Bối cảnh + Định dạng (kèm ví dụ).\n"
     "3. Hỏi tiếp (follow-up): chỉnh dần — “viết ngắn hơn”, “trang trọng hơn”.",
     "Viết lại công việc Buổi 1 bằng công thức 4 phần; so sánh trước/sau.",
     "Áp dụng công thức cho 2 việc khác trong tuần."),
    ("Buổi 3\nĐưa AI vào công việc thật",
     "Dùng AI để làm chính công việc của mình nhanh hơn.",
     "1. Đọc & tóm tắt tài liệu bằng AI: tải file (PDF, Word) lên rồi hỏi.\n"
     "2. Công cụ 2026 không chỉ là khung chat: tải tệp lên, lưu ngữ cảnh lâu dài (Projects / Custom GPTs).\n"
     "3. Luôn kiểm tra lại (kỹ năng riêng): bảo AI trích nguồn; tự hỏi “có dám lấy tên mình bảo đảm không?”.",
     "Bốc 1 công việc thật trong tuần, dùng AI làm nhanh hơn (kết hợp KB nội bộ nếu có).",
     "Tìm 1 việc lặp đi lặp lại → giao hẳn cho AI."),
    ("Buổi 4\nTối ưu & Biến thành thói quen",
     "Giữ thói quen lâu dài, không “học xong rồi bỏ”.",
     "1. Lưu prompt hay (Prompt Templates).\n"
     "2. Xây thư viện prompt chung của nhóm (Google Doc / Notion).\n"
     "3. Chia sẻ kết quả thật → tạo động lực cho người khác dùng.\n"
     "4. Chốt thói quen: mỗi người nói rõ 1 việc đã giao hẳn cho AI.\n"
     "5. Giải đáp vướng mắc.",
     "Rà lại 3 chỉ số thành công ở đầu khóa.",
     "Duy trì công việc đã giao cho AI."),
]
for i, s in enumerate(sessions):
    r = data_row(ws, s, r, zebra=(i % 2 == 1))
r = spacer(r)

r = section(ws, "BỎ TÚI NHANH", 5, r)
r = kv(ws, "Công thức prompt", "Vai trò + Nhiệm vụ + Bối cảnh + Định dạng", r)
r = banner(ws,
    "QUY TẮC BẢO MẬT — Không dán vào AI: mật khẩu, dữ liệu khách hàng, thông tin nội bộ nhạy cảm.",
    5, r, color=WARN)

ws.freeze_panes = "A2"

# ============================================================== SHEET 2
ws2 = wb.create_sheet("Kế hoạch KB")
widths2 = [16, 30, 50, 30]
for i, w in enumerate(widths2, start=1):
    ws2.column_dimensions[chr(64 + i)].width = w

r = 1
r = title(ws2, "KẾ HOẠCH XÂY DỰNG KNOWLEDGE BASE (NotebookLM)", 4, r)
r = kv(ws2, "Mục tiêu", "Tạo “bộ não chung” cho team — hỏi bằng ngôn ngữ tự nhiên, nhận câu trả lời CÓ trích dẫn nguồn.", r)
r = spacer(r)

r = banner(ws2,
    "⚠️ ĐỌC TRƯỚC KHI LÀM: NotebookLM CHỈ tự đồng bộ với file Google gốc (Docs, Sheets, Slides). "
    "File PDF / .md / dán tay KHÔNG tự cập nhật → dễ trả lời theo bản cũ.\n"
    "→ BẮT BUỘC: tài liệu hay thay đổi (Manual, Monitoring, Incident) phải viết dưới dạng Google Docs/Sheets.",
    4, r, color=WARN)
r = spacer(r)

r = section(ws2, "4 KNOWLEDGE BASE (chia theo vòng đời tài liệu)", 4, r)
r = header_row(ws2, ["KB", "Loại tài liệu", "Đặc điểm", ""], r)
ws2.merge_cells(start_row=r - 1, start_column=3, end_row=r - 1, end_column=4)
kbs = [
    ("01_Onboarding", "Giới thiệu, hướng dẫn nhập môn", "Đọc 1 lần, ít đổi"),
    ("02_Manual", "Hướng dẫn vận hành, runbook", "Tham khảo thường xuyên"),
    ("03_Incident", "Nhật ký sự cố", "Chỉ thêm mới, lớn nhanh"),
    ("04_Monitoring", "Giám sát, dashboard", "Tham khảo, có thay đổi"),
]
for i, k in enumerate(kbs):
    ws2.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    r = data_row(ws2, [k[0], k[1], k[2]], r, zebra=(i % 2 == 1))
    ws2.cell(row=r - 1, column=4).border = BORDER
r = spacer(r)

r = section(ws2, "CÁC BƯỚC THỰC HIỆN", 4, r)
r = header_row(ws2, ["Bước", "Nội dung", "", ""], r)
ws2.merge_cells(start_row=r - 1, start_column=2, end_row=r - 1, end_column=4)
steps = [
    ("Bước 1\nTạo thư mục Drive",
     "Tạo cây thư mục: AI_KB/ (thư mục cha) chứa 01_Onboarding, 02_Manual, 03_Incident, 04_Monitoring."),
    ("Bước 2\nChuẩn bị tài liệu",
     "Viết tài liệu dạng Google Docs/Sheets (để tự đồng bộ).\n"
     "• Onboarding: Org_Overview, Security_Training\n"
     "• Manual: Chrome_Vup_Test, RUNBOOK_Deploy_ServiceA\n"
     "• Incident: Incident_Template + 1 Google Sheet TỔNG ghi mọi sự cố (ĐỪNG tạo mỗi sự cố 1 file → tránh chạm giới hạn). Quy ước: INC_YYYY_###_TenSuCo\n"
     "• Monitoring: Monitoring_Overview, Dashboard_List"),
    ("Bước 3\nTạo NotebookLM",
     "Tạo 4 NotebookLM, mỗi cái nạp nguồn từ 1 thư mục KB.\n"
     "Lưu ý: notebook KHÔNG hỏi chéo nhau. Bắt đầu với 4; khi mọi người bối rối “không biết hỏi notebook nào” → gộp lại thành 1.\n"
     "Giới hạn: Free 50 nguồn/notebook; Plus 100."),
    ("Bước 4\nCâu hỏi click sẵn",
     "NotebookLM không có nút câu hỏi cố định → tạo 1 Note tên “Câu hỏi thường gặp” ghim đầu mỗi notebook; người dùng copy-paste để hỏi (vừa là mẫu, vừa là hướng dẫn)."),
    ("Bước 5\nKiểm tra (bắt buộc)",
     "Chạy thử toàn bộ câu hỏi thường gặp trước khi phát hành. Xác nhận: (1) trả lời ĐÚNG; (2) có trích dẫn ĐÚNG tài liệu. Trong vận hành, trả lời sai mà tự tin = một sự cố."),
]
for i, s in enumerate(steps):
    ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r = data_row(ws2, [s[0], s[1]], r, zebra=(i % 2 == 1))
    for col in (3, 4):
        ws2.cell(row=r - 1, column=col).border = BORDER
r = spacer(r)

r = section(ws2, "PHÂN CÔNG NGƯỜI PHỤ TRÁCH (KB không ai giữ sẽ mục nát)", 4, r)
r = header_row(ws2, ["KB", "Người phụ trách", "Việc định kỳ", ""], r)
ws2.merge_cells(start_row=r - 1, start_column=3, end_row=r - 1, end_column=4)
owners = [
    ("01_Onboarding", "", "Cập nhật khi quy trình đổi"),
    ("02_Manual", "", "Cập nhật khi runbook đổi"),
    ("03_Incident", "", "Mỗi sự cố → ghi vào Sheet"),
    ("04_Monitoring", "", "Cập nhật khi dashboard đổi"),
]
for i, o in enumerate(owners):
    ws2.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    r = data_row(ws2, [o[0], o[1], o[2]], r, zebra=(i % 2 == 1))
    ws2.cell(row=r - 1, column=4).border = BORDER
r = spacer(r)

r = section(ws2, "BẢO MẬT & PHÂN QUYỀN", 4, r)
r = banner(ws2,
    "NotebookLM tuân theo quyền của thư mục Google Drive → QUYỀN THƯ MỤC DRIVE chính là lớp kiểm soát truy cập. "
    "Khóa chặt thư mục 03_Incident và file Security_Training cho đúng người.",
    4, r, color=LIGHT)
r = spacer(r)

r = section(ws2, "NotebookLM vs. GEMINI TRONG DRIVE (dùng bổ sung)", 4, r)
r = header_row(ws2, ["Công cụ", "Dùng khi nào", "Đặc điểm", ""], r)
ws2.merge_cells(start_row=r - 1, start_column=3, end_row=r - 1, end_column=4)
tools = [
    ("Gemini trong Drive", "“File đó nằm ở đâu?”", "Tìm khắp Drive, rộng, ít trích dẫn"),
    ("NotebookLM", "“Runbook nói chính xác gì?”", "Phạm vi gọn, trích dẫn rõ → ưu tiên cho KB"),
]
for i, t in enumerate(tools):
    ws2.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    r = data_row(ws2, [t[0], t[1], t[2]], r, zebra=(i % 2 == 1))
    ws2.cell(row=r - 1, column=4).border = BORDER

ws2.freeze_panes = "A2"

out = "/Users/hit/ai-training/Khoa-hoc-AI-va-Ke-hoach-KB.xlsx"
wb.save(out)
print("saved:", out)
